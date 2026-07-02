import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook.defined_name import DefinedName
import logging

log = logging.getLogger(__name__)

class BudgetExcelBuilder:
    """Generates a four-sheet Excel workbook for the Budget Breakdown."""
    
    def __init__(self, spec):
        self.config = spec.budget.model_dump()
        self.wb = Workbook()
        self.categories = self.config['categories']
        self.line_items = self.config['line_items']
        self.burn_rate = self.config['monthly_burn_rate']
        self.styling = self.config.get('styling', {})
        
        # Computed totals
        self.total_budget = sum(c['budget'] for c in self.categories)
        self.total_actual = sum(c.get('actual', 0) for c in self.categories if c.get('actual') is not None)
        self.total_variance = self.total_actual - self.total_budget

    def _style_header_row(self, ws, row: int, cols: int) -> None:
        """Apply navy blue header styling to a row."""
        bg_color = self.styling.get("header_fill", "#1a237e").strip("#")
        fg_color = self.styling.get("header_text", "#FFFFFF").strip("#")
        header_fill = PatternFill("solid", fgColor=bg_color)
        header_font = Font(color=fg_color, bold=True, size=9)
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _apply_currency_format(self, cell) -> None:
        cell.number_format = '"$"#,##0'

    def _apply_percentage_format(self, cell) -> None:
        cell.number_format = '0.0%'

    def build_summary_sheet(self) -> None:
        ws = self.wb.active
        ws.title = "Budget Summary"
        
        # ── Title Block ──
        ws['A1'] = "BUDGET BREAKDOWN"
        ws['A1'].font = Font(bold=True, size=14, color=self.styling.get("header_fill", "#1a237e").strip("#"))
        ws.merge_cells('A1:D1')
        
        ws['A2'] = "Project:"
        ws['B2'] = self.config.get('project_name', '')
        ws['A3'] = "Date:"
        ws['B3'] = self.config.get('date', '')
        ws['A4'] = "Currency:"
        ws['B4'] = self.config.get('currency', 'USD')
        
        # ── Category Summary Header ──
        header_row = 6
        ws.cell(header_row, 1, "CATEGORY")
        ws.cell(header_row, 2, "BUDGET")
        ws.cell(header_row, 3, "PERCENTAGE")
        ws.cell(header_row, 4, "NOTES")
        self._style_header_row(ws, header_row, 4)
        
        data_start = header_row + 1
        for idx, cat in enumerate(self.categories):
            row = data_start + idx
            ws.cell(row, 1, cat['name'])
            ws.cell(row, 2, cat['budget'])
            self._apply_currency_format(ws.cell(row, 2))
            ws.cell(row, 3, cat['budget'] / self.total_budget if self.total_budget else 0)
            self._apply_percentage_format(ws.cell(row, 3))
            ws.cell(row, 4, cat.get('notes', ''))
            
            if idx % 2 == 0:
                fill = PatternFill("solid", fgColor=self.styling.get("alt_row_fill", "#F5F5F5").strip("#"))
                for col in range(1, 5):
                    ws.cell(row, col).fill = fill
                    
        total_row = data_start + len(self.categories)
        ws.cell(total_row, 1, "TOTAL BUDGET")
        ws.cell(total_row, 2, self.total_budget)
        ws.cell(total_row, 3, 1.0)
        total_fill = PatternFill("solid", fgColor=self.styling.get("total_row_fill", "#E3F2FD").strip("#"))
        total_font = Font(bold=True)
        for col in range(1, 5):
            ws.cell(total_row, col).fill = total_fill
            ws.cell(total_row, col).font = total_font
        self._apply_currency_format(ws.cell(total_row, 2))
        self._apply_percentage_format(ws.cell(total_row, 3))
        
        # ── Budget vs Actual Section ──
        bva_start = total_row + 2
        ws.cell(bva_start, 1, "BUDGET VS ACTUAL")
        ws.cell(bva_start, 1).font = Font(bold=True, size=10, color=self.styling.get("header_fill", "#1a237e").strip("#"))
        
        bva_header = bva_start + 1
        for col, title in enumerate(["Category", "Budget", "Actual", "Variance"], 1):
            ws.cell(bva_header, col, title)
        self._style_header_row(ws, bva_header, 4)
        
        for idx, cat in enumerate(self.categories):
            row = bva_header + 1 + idx
            ws.cell(row, 1, cat['name'])
            ws.cell(row, 2, cat['budget'])
            ws.cell(row, 3, cat.get('actual', 0) if cat.get('actual') is not None else 0)
            variance = ws.cell(row, 3).value - ws.cell(row, 2).value
            ws.cell(row, 4, variance)
            
            for col in (2, 3, 4):
                self._apply_currency_format(ws.cell(row, col))
            
            var_cell = ws.cell(row, 4)
            if variance < 0:
                var_cell.font = Font(color=self.styling.get("positive_variance", "#4CAF50").strip("#"), bold=True)
            elif variance > 0:
                var_cell.font = Font(color=self.styling.get("negative_variance", "#E53935").strip("#"), bold=True)
        
        bva_total = bva_header + 1 + len(self.categories)
        ws.cell(bva_total, 1, "TOTAL")
        ws.cell(bva_total, 2, self.total_budget)
        ws.cell(bva_total, 3, self.total_actual)
        ws.cell(bva_total, 4, self.total_variance)
        for col in (2, 3, 4):
            self._apply_currency_format(ws.cell(bva_total, col))
        for col in range(1, 5):
            ws.cell(bva_total, col).fill = total_fill
            ws.cell(bva_total, col).font = Font(bold=True)
            
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 40

    def build_detail_sheet(self) -> None:
        ws = self.wb.create_sheet("Detailed Breakdown")
        headers = ["CATEGORY", "ITEM DESCRIPTION", "QTY", "UNIT COST", "TOTAL"]
        for col, h in enumerate(headers, 1):
            ws.cell(1, col, h)
        self._style_header_row(ws, 1, len(headers))
        
        for idx, item in enumerate(self.line_items):
            row = idx + 2
            ws.cell(row, 1, item['category'])
            ws.cell(row, 2, item['item'])
            ws.cell(row, 3, item.get('qty', 1))
            ws.cell(row, 4, item.get('unit_cost', 0))
            self._apply_currency_format(ws.cell(row, 4))
            ws.cell(row, 5, f"=C{row}*D{row}")
            self._apply_currency_format(ws.cell(row, 5))
            
            if idx % 2 == 0:
                fill = PatternFill("solid", fgColor=self.styling.get("alt_row_fill", "#F5F5F5").strip("#"))
                for col in range(1, 6):
                    ws.cell(row, col).fill = fill
                    
        total_row = len(self.line_items) + 2
        ws.cell(total_row, 1, "TOTAL")
        ws.cell(total_row, 5, f"=SUM(E2:E{total_row - 1})")
        self._apply_currency_format(ws.cell(total_row, 5))
        total_fill = PatternFill("solid", fgColor=self.styling.get("total_row_fill", "#E3F2FD").strip("#"))
        for col in range(1, 6):
            ws.cell(total_row, col).fill = total_fill
            ws.cell(total_row, col).font = Font(bold=True)
            
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16

    def build_burn_rate_sheet(self) -> None:
        ws = self.wb.create_sheet("Monthly Burn Rate")
        headers = ["MONTH", "PLANNED", "ACTUAL", "CUM. PLANNED", "CUM. ACTUAL"]
        for col, h in enumerate(headers, 1):
            ws.cell(1, col, h)
        self._style_header_row(ws, 1, len(headers))
        
        for idx, month_data in enumerate(self.burn_rate):
            row = idx + 2
            ws.cell(row, 1, month_data['month'])
            ws.cell(row, 2, month_data['planned'])
            actual = month_data.get('actual')
            if actual is not None:
                ws.cell(row, 3, actual)
                self._apply_currency_format(ws.cell(row, 3))
            self._apply_currency_format(ws.cell(row, 2))
            
            if row == 2:
                ws.cell(row, 4, f"=B{row}")
                ws.cell(row, 5, f"=C{row}" if actual is not None else "")
            else:
                ws.cell(row, 4, f"=D{row - 1}+B{row}")
                ws.cell(row, 5, f"=E{row - 1}+C{row}" if actual is not None else "")
                
            for col in (4, 5):
                self._apply_currency_format(ws.cell(row, col))
                
            if idx % 2 == 0:
                fill = PatternFill("solid", fgColor=self.styling.get("alt_row_fill", "#F5F5F5").strip("#"))
                for col in range(1, 6):
                    ws.cell(row, col).fill = fill
                    
        ws.column_dimensions['A'].width = 14
        for col_letter in ('B', 'C', 'D', 'E'):
            ws.column_dimensions[col_letter].width = 16

    def build_data_connection_sheet(self) -> None:
        ws = self.wb.create_sheet("DataConnection")
        headers = ["VISIO_ITEM", "VALUE", "CATEGORY", "NOTE"]
        for col, h in enumerate(headers, 1):
            ws.cell(1, col, h)
        self._style_header_row(ws, 1, len(headers))
        
        rows = [
            ("TotalBudget", self.total_budget, "Summary", "Linked to Visio KPI TotalBudget box"),
            ("TotalActual", self.total_actual, "Summary", "Linked to Visio KPI ActualSpend box"),
            ("Remaining", self.total_budget - self.total_actual, "Summary", "Linked to Visio KPI Remaining box"),
        ]
        
        for cat in self.categories:
            rows.append((cat['name'], cat['budget'], "Category", f"Linked to Visio bar chart: {cat['name']} bar width"))
            
        for cat in self.categories:
            pct = round(cat['budget'] / self.total_budget * 100, 1) if self.total_budget else 0
            rows.append((f"{cat['name']}Pct", pct, "Percentage", f"Linked to Visio pie sector angle: {cat['name']}"))
            
        for idx, (item, value, category, note) in enumerate(rows):
            row = idx + 2
            ws.cell(row, 1, item)
            ws.cell(row, 2, value)
            ws.cell(row, 3, category)
            ws.cell(row, 4, note)
            
            if idx % 2 == 0:
                fill = PatternFill("solid", fgColor=self.styling.get("alt_row_fill", "#F5F5F5").strip("#"))
                for col in range(1, 5):
                    ws.cell(row, col).fill = fill
                    
        for idx, (item, _, _, _) in enumerate(rows):
            safe = re.sub(r"[^\w]", "_", str(item))
            if not safe or safe[0].isdigit():
                safe = f"Item_{idx}_{safe}"
            cell_ref = f"DataConnection!$B${idx + 2}"
            self.wb.defined_names.add(DefinedName(name=safe[:255], attr_text=cell_ref))
            
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 50

    def build(self) -> None:
        self.build_summary_sheet()
        self.build_detail_sheet()
        self.build_burn_rate_sheet()
        self.build_data_connection_sheet()

    def save(self, output_path: str) -> None:
        self.wb.save(output_path)
        log.info(f"Excel workbook saved to {output_path}")
